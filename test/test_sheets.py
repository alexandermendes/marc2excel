# -*- coding: utf8 -*-

import openpyxl
import pytest
from marc2excel.exceptions import InvalidMARCCodeError
from marc2excel.sheets import Sheets

sheets = Sheets(silent=True)


class TestSheets():

    def create_workbook(self, tmp_xlsx, data):
        wb = openpyxl.Workbook(write_only=True)
        ws = wb.create_sheet()
        for row in data:
            ws.append(row)
        wb.save(tmp_xlsx)

    def test_write(self, tmp_xlsx):
        data = [['h1', 'h2'], ['v1', 'v2']]
        sheets.write(tmp_xlsx, data)
        wb = openpyxl.load_workbook(tmp_xlsx)
        ws = wb.get_active_sheet()
        assert [ws['A1'].value, ws['B1'].value] == data[0]
        assert [ws['A2'].value, ws['B2'].value] == data[1]
        assert ws.max_row == 2

    def test_extract_valid_data(self, tmp_xlsx):
        data = [['ldr', '001', '245\\\\$a', '245\\\\$a[2]'],
                ['0', '123', 't1', 't2'], ['0', '456', 't3', 't4']]
        self.create_workbook(tmp_xlsx, data)
        extracted = sheets.extract_data(tmp_xlsx, 0)
        ret = [{u'ldr': u'0', u'001': u'123', u'245\\\\$a': u't1',
                u'245\\\\$a[2]': u't2'},
               {u'ldr': u'0', u'001': u'456', u'245\\\\$a': u't3',
                u'245\\\\$a[2]': u't4'}]
        assert extracted == ret

    def test_extract_invalid_field_code(self, tmp_xlsx):
        h = '24\\$a'
        data = [[h], ['t1']]
        self.create_workbook(tmp_xlsx, data)
        with pytest.raises(InvalidMARCCodeError) as excinfo:
            extracted = sheets.extract_data(tmp_xlsx, 0)
        assert h in str(excinfo.value)

    def test_extract_invalid_indicator(self, tmp_xlsx):
        h = '245xx$a'
        data = [[h], ['t1']]
        self.create_workbook(tmp_xlsx, data)
        with pytest.raises(InvalidMARCCodeError) as excinfo:
            extracted = sheets.extract_data(tmp_xlsx, 0)
        assert h in str(excinfo.value)

    def test_extract_invalid_subfield_code(self, tmp_xlsx):
        h = '245\\$aa'
        data = [[h], ['t1']]
        self.create_workbook(tmp_xlsx, data)
        with pytest.raises(InvalidMARCCodeError) as excinfo:
            extracted = sheets.extract_data(tmp_xlsx, 0)
        assert h in str(excinfo.value)

    def test_extract_invalid_repeated_field_code(self, tmp_xlsx):
        h = '245\\$aa[a]'
        data = [[h], ['t1']]
        self.create_workbook(tmp_xlsx, data)
        with pytest.raises(InvalidMARCCodeError) as excinfo:
            extracted = sheets.extract_data(tmp_xlsx, 0)
        assert h in str(excinfo.value)
