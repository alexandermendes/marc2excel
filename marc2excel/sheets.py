# -*- coding: utf-8 -*-
"""Sheets module for marc2excel."""

import re
import openpyxl
from tqdm import tqdm
from marc2excel.exceptions import InvalidMARCCodeError


class Sheets(object):
    """A class for reading and writing Excel spreadsheets.

    :param silent: Don't display CLI progress.
    """

    def __init__(self, silent=True):
        self.silent = silent

    def write(self, path, data):
        """Write data to an Excel spreadsheet.

        :param path: Path to save the spreadsheet.
        :param data: The data as a list of lists of strings.
        """
        wb = openpyxl.Workbook(write_only=True)
        ws = wb.create_sheet()

        for row in tqdm(data, unit="rows", desc="Writing Excel data",
                        disable=self.silent):
            ws.append(row)
        wb.save(path)

    def extract_data(self, path, sheet):
        """Return all data extracted from an Excel spreadsheet.

        :param path: Path to the Excel spreadsheet.
        :param sheet: The index of the sheet from which to extract data.

        :returns: The MARC data as a list of dicts.
        """
        wb = openpyxl.load_workbook(path, read_only=True)
        ws = wb.worksheets[sheet]

        headers = []
        data = []
        for row in tqdm(ws.iter_rows(), unit="rows",
                        desc="Extracting Excel data", total=ws.max_row,
                        disable=self.silent):

            if not headers:
                headers = [cell.value for cell in row if cell.value]
                self._validate_headers(headers, path)
                continue

            record = {headers[cell.column - 1]: cell.value
                      for cell in row if cell.value}
            data.append(record)
        return data

    def _validate_headers(self, headers, path):
        """Validate headers extracted from an Excel spreadsheet."""
        ptn = r'^(0(0[1-9]|10)|ldr|[0-9]{3}[\\0-9]{2}\$[0-9a-z])(\[[1-9]+\])?$'
        marc_re = re.compile(ptn)
        for h in headers:
            key = re.sub(r'\s+', '', h.lower())
            if not marc_re.match(key):
                raise InvalidMARCCodeError(h, path)
